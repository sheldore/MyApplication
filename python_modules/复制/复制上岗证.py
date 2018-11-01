from openpyxl import load_workbook
from openpyxl.styles import numbers,Font,colors

对内表book = load_workbook("燕子山矿井下从业人员培训持证上岗情况lixia.xlsm",keep_vba=True)
对内表单位目录sheet = 对内表book["单位目录"]
上岗证sheet = load_workbook("合格证上岗证2018.xlsm").active
上岗证办证时间 = 上岗证sheet.cell(2,6).value.replace("时间","").replace(" ","").replace(":","").replace("：","")
单位转换dict = {"温州建设项目部":"温州建设","宏泰项目部":"宏泰三部","机电科":"机电部","财务科":"混编单位","宏远安装队":"宏远公司","地测科":"地质队"}
success = 0
fail = 0

for 上岗证row in 上岗证sheet["b5:i"+str(上岗证sheet.max_row)]:
    上岗证某单位cell,上岗证某姓名cell,_,_,上岗证工种cell,_,_,上岗证某身份证cell = 上岗证row
    对内表单位list = [对内表单位目录sheet.cell(i,2).value for i in range(4,对内表单位目录sheet.max_row)]#['综采一队','综采二队',...]
    if 上岗证某单位cell.value not in 对内表单位list:
        上岗证某单位cell.value =  单位转换dict[上岗证某单位cell.value]
    if 上岗证某单位cell.value in 对内表单位list:
        对内表某单位sheet = 对内表book[上岗证某单位cell.value]
        对内表某单位sheet姓名list = [对内表某单位sheet.cell(i,3).value for i in range(5,对内表某单位sheet.max_row+1)]
        对内表某单位sheet身份证list = [对内表某单位sheet.cell(i,10).value[2:] if 对内表某单位sheet.cell(i,10).value is not None else "" for i in range(5,对内表某单位sheet.max_row+1)]
        #如果身份证号相同
        if 上岗证某身份证cell.value in  对内表某单位sheet身份证list:
            此身份证的行号= 对内表某单位sheet身份证list.index(上岗证某身份证cell.value)+5
            #如果上岗证工种无记录或者记录颜色不为红色则覆盖
            if 对内表某单位sheet.cell(此身份证的行号,8).value == None or \
                    len(str(对内表某单位sheet.cell(此身份证的行号,8).value).replace(" ","")) == 0 \
                    or 对内表某单位sheet.cell(此身份证的行号,8).font.color.value != colors.RED:
                对内表某单位sheet.cell(此身份证的行号,7).value = 上岗证办证时间
                对内表某单位sheet.cell(此身份证的行号,8).value = 上岗证工种cell.value
                对内表某单位sheet.cell(此身份证的行号,10).value = str(12)+上岗证某身份证cell.value
                对内表某单位sheet.cell(此身份证的行号,7).font = Font(color=colors.RED)
                对内表某单位sheet.cell(此身份证的行号,8).font = Font(color=colors.RED)
                success +=1
                print(上岗证某单位cell.value +" " +上岗证某姓名cell.value + " " + 上岗证某身份证cell.value + "身份证号相同, 无上岗证工种记录,添加成功" )
            #如果上岗证工种有记录且为红色,对内表中工种记录与上岗证上岗证办证记录不一致,则新增一行对内表某单位sheet.insert_rows(此身份证的行号)
            elif 对内表某单位sheet.cell(此身份证的行号,7).font.color.value == colors.RED and 对内表某单位sheet.cell(此身份证的行号,8).value != 上岗证工种cell.value:
                    对内表某单位sheet.cell(此身份证的行号,7).value = 上岗证办证时间
                    对内表某单位sheet.cell(此身份证的行号,8).value = 上岗证工种cell.value
                    对内表某单位sheet.cell(此身份证的行号,10).value = str(12)+上岗证某身份证cell.value
                    对内表某单位sheet.cell(此身份证的行号,7).font = Font(color=colors.RED)
                    对内表某单位sheet.cell(此身份证的行号,8).font = Font(color=colors.RED)
                    success +=1
                    print(上岗证某单位cell.value +" " +上岗证某姓名cell.value + " " + 上岗证某身份证cell.value + "身份证号相同, 有不一致的工种记录,添加成功" )
            else:
                    fail+=1
                    print(上岗证某单位cell.value +" " +上岗证某姓名cell.value + " " + 上岗证某身份证cell.value + "身份证号相同, 因与对内表上岗证工种记录一致而未添加" )
        #如果身份证号不同,姓名相同
        elif 上岗证某姓名cell.value in 对内表某单位sheet姓名list:
            此姓名的行号= 对内表某单位sheet姓名list.index(上岗证某姓名cell.value)+5
            is_same_person = input(上岗证某单位cell.value + 上岗证某姓名cell.value+"的身份证号不同"+chr(10)\
                                        +"对内表:       "+(对内表某单位sheet.cell(此姓名的行号,10).value[2:] if 对内表某单位sheet.cell(此姓名的行号,10).value is not None else "") +chr(10)\
                                        +"上岗证上岗证表:"+上岗证某身份证cell.value+chr(10)\
                                        +"是否为同一人? 是Y,否N")
            #如果是同一人
            if  is_same_person=="y" or is_same_person=="Y":
                #如果上岗证工种无记录或者记录颜色不为红色则覆盖
                if 对内表某单位sheet.cell(此姓名的行号,8).value == None or \
                        len(str(对内表某单位sheet.cell(此姓名的行号,8).value).replace(" ","")) == 0 \
                        or 对内表某单位sheet.cell(此姓名的行号,8).font.color.value != colors.RED:
                    对内表某单位sheet.cell(此姓名的行号,7).value = 上岗证办证时间
                    对内表某单位sheet.cell(此姓名的行号,8).value = 上岗证工种cell.value
                    对内表某单位sheet.cell(此姓名的行号,10).value = str(12)+上岗证某身份证cell.value
                    对内表某单位sheet.cell(此姓名的行号,7).font = Font(color=colors.RED)
                    对内表某单位sheet.cell(此姓名的行号,8).font = Font(color=colors.RED)
                    success +=1
                    print(上岗证某单位cell.value +" " +上岗证某姓名cell.value + " " + 上岗证某身份证cell.value + "有相同的姓名,身份证号不同, 经认定是同一人,无上岗证工种记录,添加成功" )
                else:#如果上岗证工种有记录且为红色
                    #如果对内表中工种记录与上岗证上岗证办证记录不一致,则新增一行
                    if 对内表某单位sheet.cell(此姓名的行号,7).font.color.value == colors.RED and 对内表某单位sheet.cell(此姓名的行号,8).value != 上岗证工种cell.value:
                        对内表某单位sheet.insert_rows(此姓名的行号)
                        对内表某单位sheet.cell(此姓名的行号,7).value = 上岗证办证时间
                        对内表某单位sheet.cell(此姓名的行号,8).value = 上岗证工种cell.value
                        对内表某单位sheet.cell(此姓名的行号,10).value = str(12)+上岗证某身份证cell.value
                        对内表某单位sheet.cell(此姓名的行号,7).font = Font(color=colors.RED)
                        对内表某单位sheet.cell(此姓名的行号,8).font = Font(color=colors.RED)
                        success +=1
                        print(上岗证某单位cell.value +" " +上岗证某姓名cell.value + " " + 上岗证某身份证cell.value + "有相同的姓名,身份证号不同,经认定是同一人, 有不一致的工种记录,添加成功" )
                    else:
                        fail+=1
                        print(上岗证某单位cell.value +" " +上岗证某姓名cell.value + " " + 上岗证某身份证cell.value + "有相同的姓名但身份证号不同,经认定是同一人, 因与对内表上岗证工种记录一致而未添加" )
            else:#如果不是同一人
                对内表某单位sheet.cell(对内表某单位sheet.max_row+1,2).value =上岗证某单位cell.value
                对内表某单位sheet.cell(对内表某单位sheet.max_row,3).value =上岗证某姓名cell.value
                对内表某单位sheet.cell(对内表某单位sheet.max_row,7).value =上岗证办证时间
                对内表某单位sheet.cell(对内表某单位sheet.max_row,8).value =上岗证工种cell.value
                对内表某单位sheet.cell(对内表某单位sheet.max_row,7).font = Font(color=colors.RED)
                对内表某单位sheet.cell(对内表某单位sheet.max_row,8).font = Font(color=colors.RED)
                对内表某单位sheet.cell(对内表某单位sheet.max_row,10).value = str(12)+上岗证某身份证cell.value
                success+=1
                print(上岗证某单位cell.value +" " +上岗证某姓名cell.value + " " + 上岗证某身份证cell.value + " 有相同的姓名但身份证号不同,经认定不是同一人,添加成功" )

        else:#如果姓名不在对内表某单位sheet姓名list中,身份证也不在对内表某单位sheet身份证list中
            对内表某单位sheet.cell(对内表某单位sheet.max_row+1,2).value =上岗证某单位cell.value
            对内表某单位sheet.cell(对内表某单位sheet.max_row,3).value =上岗证某姓名cell.value
            对内表某单位sheet.cell(对内表某单位sheet.max_row,7).value =上岗证办证时间
            对内表某单位sheet.cell(对内表某单位sheet.max_row,8).value =上岗证工种cell.value
            对内表某单位sheet.cell(对内表某单位sheet.max_row,7).font = Font(color=colors.RED)
            对内表某单位sheet.cell(对内表某单位sheet.max_row,8).font = Font(color=colors.RED)
            对内表某单位sheet.cell(对内表某单位sheet.max_row,10).value = str(12)+上岗证某身份证cell.value
            success+=1
            print(上岗证某单位cell.value +" " +上岗证某姓名cell.value + " " + 上岗证某身份证cell.value + " 无记录的姓名,添加成功" )
print("正在保存文件...")
对内表book.save("燕子山矿井下从业人员培训持证上岗情况lixia.xlsm")
input(str(success) +"条记录添加成功"+chr(10)
        +str(fail) + "条记录添加失败"+chr(10)
        +"文件保存成功,请按任意键退出..")