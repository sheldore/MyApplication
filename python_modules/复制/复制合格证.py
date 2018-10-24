from openpyxl import load_workbook
from openpyxl.styles import numbers,Font,colors

对内表book = load_workbook("1.xlsm",keep_vba=True)
对内表单位目录sheet = 对内表book["单位目录"]
合格证sheet = load_workbook("合格证上岗证2018.xlsm").active
合格证办证时间 = 合格证sheet.cell(2,1).value.replace("时间","").replace(" ","").replace(":","").replace("：","")
单位转换dict = {"温州建设项目部":"温州建设"}

success = 0
fail = 0
for 合格证row in 合格证sheet["b5:h"+str(合格证sheet.max_row)]:
    合格证某单位cell,合格证某姓名cell,_,_,_,_,合格证某身份证cell = 合格证row
    对内表单位list = [对内表单位目录sheet.cell(i,2).value for i in range(4,对内表单位目录sheet.max_row)]#['综采一队','综采二队',...]
    if 合格证某单位cell.value not in 对内表单位list:
        合格证某单位cell.value =  单位转换dict[合格证某单位cell.value]
    if 合格证某单位cell.value in 对内表单位list:
        对内表某单位sheet = 对内表book[str(对内表单位list.index(合格证某单位cell.value)+1) + 合格证某单位cell.value]
        对内表某单位sheet姓名list = [对内表某单位sheet.cell(i,3).value for i in range(5,对内表某单位sheet.max_row+1)]
        对内表某单位sheet身份证list = [对内表某单位sheet.cell(i,10).value[2:] if 对内表某单位sheet.cell(i,10).value is not None else "" for i in range(5,对内表某单位sheet.max_row+1)]
        #如果身份证号相同
        if 合格证某身份证cell.value in  对内表某单位sheet身份证list:
            此身份证的行号= 对内表某单位sheet身份证list.index(合格证某身份证cell.value)+5
            #如果合格证初训无记录或颜色不为红
            if 对内表某单位sheet.cell(此身份证的行号,4).value is None \
                    or len(str(对内表某单位sheet.cell(此身份证的行号,4).value).replace(" ","")) == 0 \
                    or 对内表某单位sheet.cell(此身份证的行号,4).font.color.value != colors.RED:
                对内表某单位sheet.cell(此身份证的行号,4).value = 合格证办证时间
                对内表某单位sheet.cell(此身份证的行号,10).value = str(12)+合格证某身份证cell.value
                对内表某单位sheet.cell(此身份证的行号,4).font = Font(color=colors.RED)
                success +=1
                print(合格证某单位cell.value +" " +合格证某姓名cell.value + " " + 合格证某身份证cell.value + "身份证号相同, 无初训记录,添加成功" )
            #如果合格证初训有记录,如果初训记录与合格证上岗证办证记录不一致,则覆盖
            elif 对内表某单位sheet.cell(此身份证的行号,4).value != 合格证办证时间 \
                    and 对内表某单位sheet.cell(此身份证的行号,4).font.color.value == colors.RED:
                对内表某单位sheet.cell(此身份证的行号,2).value =合格证某单位cell.value
                对内表某单位sheet.cell(此身份证的行号,3).value =合格证某姓名cell.value
                对内表某单位sheet.cell(此身份证的行号,4).value =合格证办证时间
                对内表某单位sheet.cell(此身份证的行号,4).font = Font(color=colors.RED)
                对内表某单位sheet.cell(此身份证的行号,10).value = str(12)+合格证某身份证cell.value
                success +=1
                print(合格证某单位cell.value +" " +合格证某姓名cell.value + " " + 合格证某身份证cell.value + "身份证号相同, 初训记录不一致,添加成功" )
            else:
                fail+=1
                print(合格证某单位cell.value +" " +合格证某姓名cell.value + " " + 合格证某身份证cell.value + "身份证号相同, 初训记录一致,未添加" )
        #如果身份证号不同,姓名相同
        elif 合格证某姓名cell.value in  对内表某单位sheet姓名list:
            此姓名的行号 = 对内表某单位sheet姓名list.index(合格证某姓名cell.value)+5
            is_same_person = input(合格证某单位cell.value + 合格证某姓名cell.value+"的身份证号不同"+chr(10)\
                                    +"对内表:       "+(对内表某单位sheet.cell(此姓名的行号,10).value[2:] if 对内表某单位sheet.cell(此姓名的行号,10).value is not None else "") +chr(10)\
                                    +"合格证上岗证表:"+合格证某身份证cell.value+chr(10)\
                                    +"是否为同一人? 是Y,否N")
            #如果是同一人
            if  is_same_person=="y" or is_same_person=="Y":
                #如果合格证初训无记录
                if 对内表某单位sheet.cell(此姓名的行号,4).value == None \
                        or len(str(对内表某单位sheet.cell(此姓名的行号,4).value).replace(" ","")) == 0 \
                        or 对内表某单位sheet.cell(此姓名的行号,4).font.color.value != colors.RED:
                    对内表某单位sheet.cell(此姓名的行号,4).value = 合格证办证时间
                    对内表某单位sheet.cell(此姓名的行号,10).value = str(12)+合格证某身份证cell.value
                    对内表某单位sheet.cell(此姓名的行号,4).font = Font(color=colors.RED)
                    success +=1
                    print(合格证某单位cell.value +" " +合格证某姓名cell.value + " " + 合格证某身份证cell.value + "有相同的姓名但身份证号不同,经认定是同一人 无初训记录,添加成功" )
                #如果合格证初训有记录
                elif 对内表某单位sheet.cell(此姓名的行号,4).value != 合格证办证时间\
                            and 对内表某单位sheet.cell(此姓名的行号,4).font.color.value == colors.RED:
                        对内表某单位sheet.cell(此姓名的行号,2).value =合格证某单位cell.value
                        对内表某单位sheet.cell(此姓名的行号,3).value =合格证某姓名cell.value
                        对内表某单位sheet.cell(此姓名的行号,4).value =合格证办证时间
                        对内表某单位sheet.cell(此姓名的行号,4).font = Font(color=colors.RED)
                        对内表某单位sheet.cell(此姓名的行号,10).value = str(12)+合格证某身份证cell.value
                        success +=1
                        print(合格证某单位cell.value +" " +合格证某姓名cell.value + " " + 合格证某身份证cell.value + "有相同的姓名但身份证号不同,经认定是同一人, 有不一致的初训记录,添加成功" )
                else:
                    fail +=1
                    print(合格证某单位cell.value +" " +合格证某姓名cell.value + " " + 合格证某身份证cell.value + "有相同的姓名但身份证号不同,经认定是同一人 因与对内表合格证初训记录一致而未添加" )
            #如果不是同一人
            else:
                对内表某单位sheet.cell(对内表某单位sheet.max_row+1,2).value =合格证某单位cell.value
                对内表某单位sheet.cell(对内表某单位sheet.max_row,3).value =合格证某姓名cell.value
                对内表某单位sheet.cell(对内表某单位sheet.max_row,4).value =合格证办证时间
                对内表某单位sheet.cell(对内表某单位sheet.max_row,4).font = Font(color=colors.RED)
                对内表某单位sheet.cell(对内表某单位sheet.max_row,10).value = str(12)+合格证某身份证cell.value
                success+=1
                print(合格证某单位cell.value +" " +合格证某姓名cell.value + " " + 合格证某身份证cell.value + " 有相同的姓名但身份证号不同,经认定不是同一人,添加成功" )

        else:#如果姓名不在对内表某单位sheet姓名list中,身份证也不在对内表某单位sheet身份证list中
            对内表某单位sheet.cell(对内表某单位sheet.max_row+1,2).value =合格证某单位cell.value
            对内表某单位sheet.cell(对内表某单位sheet.max_row,3).value =合格证某姓名cell.value
            对内表某单位sheet.cell(对内表某单位sheet.max_row,4).value =合格证办证时间
            对内表某单位sheet.cell(对内表某单位sheet.max_row,4).font = Font(color=colors.RED)
            对内表某单位sheet.cell(对内表某单位sheet.max_row,10).value = str(12)+合格证某身份证cell.value
            success+=1
            print(合格证某单位cell.value +" " +合格证某姓名cell.value + " " + 合格证某身份证cell.value + " 新增的姓名,添加成功" )

print("正在保存文件...")
对内表book.save("1.xlsm")
input(str(success) +"条记录添加成功"+chr(10)
        +str(fail) + "条记录添加失败"+chr(10)
        +"文件保存成功,请按任意键退出..")