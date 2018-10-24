from openpyxl import Workbook,load_workbook

单位名称 = "技术部"
井下单位book = load_workbook("井下单位Excel工作表赵芳.xlsm")
持证book = load_workbook("燕子山矿井下从业人员培训持证上岗情况lixia.xlsm")
在册人员名单book = load_workbook("各单位在册人员名单.xlsx")
在册核对结果book = Workbook()

井下某单位sheet = 井下单位book[单位名称]
在册人员名单某单位sheet = 在册人员名单book[单位名称]
在册核对结果某单位sheet = 在册核对结果book.create_sheet(title=单位名称)

持证表sheetnamelist = 持证book.sheetnames
持证表sheetnamelist.remove("单位目录")
持证表sheetnamelist.remove("0矿办")

持证表所有单位姓名list = [["".join(list(filter(lambda x:x.isalpha(),item))),
                  持证book[item].cell(i,3).value,
                  "".join(list(filter(lambda x:x.isalpha(),item))) if 持证book[item].cell(i,4).value is not None and str(持证book[item].cell(i,4).value).replace(" ","") != "" else "",
                  持证book[item].cell(i,8).value if 持证book[item].cell(i,8).value is not None and str(持证book[item].cell(i,8).value).replace(" ","") != "" else 持证book[item].cell(i,9).value if 持证book[item].cell(i,9).value is not None else "",
                  持证book[item].cell(i,10).value[2:] if 持证book[item].cell(i,10).value is not None and str(持证book[item].cell(i,10).value).replace(" ","") != "" else ""
                  ]
                 for item in 持证表sheetnamelist for i in range(5,持证book[item].max_row)]
                #['综采一队', '陈伟', '合格证所在单位', '井下单体支护工', '140211198212234739']

井下某单位姓名list = [[井下某单位sheet.cell(i,1).value,
                井下某单位sheet.cell(i,3).value if 井下某单位sheet.cell(i,3).value is None else 井下某单位sheet.cell(i,3).value.replace(" ",""),
                井下某单位sheet.cell(i,5).value if 井下某单位sheet.cell(i,5).value is None else 井下某单位sheet.cell(i,5).value.replace(" ","")]
               for i in range(2,井下某单位sheet.max_row+1) if 井下某单位sheet.cell(i,1).value is not None and 井下某单位sheet.cell(i,1).value != 井下某单位sheet.cell(i-1,1).value]
                #[144, '杨建龙', '140203196806212335']

井下单位sheetnamelist = 井下单位book.sheetnames
井下单位sheetnamelist.remove('单位目录')
井下单位sheetnamelist.remove('模板')
井下单位sheetnamelist.remove('Sheet1')
井下所有单位姓名list = [[井下单位book[j].cell(i, 1).value,
                井下单位book[j].cell(i, 2).value if 井下单位book[j].cell(i, 2).value is None else str(井下单位book[j].cell(i, 2).value).replace(" ", ""),
                井下单位book[j].cell(i, 3).value if 井下单位book[j].cell(i, 3).value is None else str(井下单位book[j].cell(i, 3).value).replace(" ", ""),
                井下单位book[j].cell(i, 5).value if 井下单位book[j].cell(i, 5).value is None else str(井下单位book[j].cell(i,5).value).replace(" ", "")]
               for j in 井下单位sheetnamelist for i in range(2, 井下单位book[j].max_row + 1)
               if  井下单位book[j].cell(i, 1).value is not None and 井下单位book[j].cell(i, 1).value != 井下单位book[j].cell(i - 1,1).value]
                #[2, '综采一队', '吴雪峰', '142225197810234785']

在册人员名单某单位姓名list = [["",#序号
                    在册人员名单某单位sheet.cell(i,4).value.replace(" ",""),#姓名
                    在册人员名单某单位sheet.cell(i,7).value.replace(" ",""),#身份证
                    在册人员名单某单位sheet.cell(i,8).value.replace(" ",""),]#职务工种
                   for i in range(4,在册人员名单某单位sheet.max_row+1)]
                    #[30, '王建伟', '140203198301044718', '综采维修电工']

for i in range(0,len(井下某单位姓名list)):
    for j in range(0,len(在册人员名单某单位姓名list)):
        if 井下某单位姓名list[i][1] == 在册人员名单某单位姓名list[j][1]:#如果姓名相同
            if [井下某单位姓名list[k][1] for k in  range(0,len(井下某单位姓名list))].count(井下某单位姓名list[i][1])==1:#如果档案姓名唯一，则直接赋档案号
                在册人员名单某单位姓名list[j][0] = 井下某单位姓名list[i][0]
            elif 井下某单位姓名list[i][2] is None or 井下某单位姓名list[i][2] == "":#如果档案姓名不唯一且档案身份证为空，则直接赋档案号
                在册人员名单某单位姓名list[j][0] = 井下某单位姓名list[i][0]
            elif 井下某单位姓名list[i][2][6:12] == 在册人员名单某单位姓名list[j][2][6:12]:#如果档案姓名不唯一且姓名与身份证都相同，则赋档案号
                在册人员名单某单位姓名list[j][0] = 井下某单位姓名list[i][0]
        elif 井下某单位姓名list[i][2] is not None and 井下某单位姓名list[i][2] != "" \
                and 井下某单位姓名list[j][2] is not None and 井下某单位姓名list[j][2] != ""\
                and 井下某单位姓名list[i][2] == 在册人员名单某单位姓名list[j][2]:#如果姓名不同,且两身份证都不空,身份证相同，则直接赋档案号
            在册人员名单某单位姓名list[j][0] = 井下某单位姓名list[i][0]

在册核对结果某单位sheet.append(["单位: ",单位名称,"档案名单","","在册名单"])
在册核对结果某单位sheet.append(["序号","姓名","身份证","序号","姓名","身份证","档案所在单位","合格证所在单位","持证表上岗证（单位及工种）","在册表上岗证（职务及工种）"])
for i in range(0,len(井下某单位姓名list)):
    在册核对结果某单位sheet.cell(i+3,1).value = 井下某单位姓名list[i][0]#档案序号
    在册核对结果某单位sheet.cell(i+3,4).value = 井下某单位姓名list[i][0]
    在册核对结果某单位sheet.cell(i+3,2).value = 井下某单位姓名list[i][1]#档案姓名
    在册核对结果某单位sheet.cell(i+3,3).value = 井下某单位姓名list[i][2]#档案身份证
    for j in range(0,len(在册人员名单某单位姓名list)):
        if 在册人员名单某单位姓名list[j][0] == 井下某单位姓名list[i][0]:#如果在册序号与档案序号相同
            在册核对结果某单位sheet.cell(i+3,5).value = 在册人员名单某单位姓名list[j][1]#在册姓名
            在册核对结果某单位sheet.cell(i+3,6).value = 在册人员名单某单位姓名list[j][2]#在册身份证
            在册核对结果某单位sheet.cell(i+3,10).value = 在册人员名单某单位姓名list[j][3]#在册职务
            for item in 持证表所有单位姓名list:
                if 在册人员名单某单位姓名list[j][2] in item:#item =  ['综采一队', '陈伟', '合格证所在单位', '井下单体支护工', '140211198212234739']如果身份证在item中
                    在册核对结果某单位sheet.cell(i+3,8).value = item[2]#复制合格证所在单位
                    在册核对结果某单位sheet.cell(i+3,9).value = (item[3] if item[2] !="" else item[0] + item[3]) if item[3] != "" else ""#如果工种不为空则复制工种
                    break
                elif 在册人员名单某单位姓名list[j][1] in item:#如果身份证不在item中，而姓名在，
                    if  在册核对结果某单位sheet.cell(i+3,8).value is None:#如果核对表中合格证所在单位为空
                        在册核对结果某单位sheet.cell(i+3,8).value = (item[2] + item[4] + " " + chr(10)) if item[2] != "" else ""#复制合格证所在单位
                    else:#如果核对表中合格证所在单位不为空
                        在册核对结果某单位sheet.cell(i+3,8).value += (item[2] + item[4] + " " + chr(10)) if item[2] != "" else ""#复制合格证所在单位
                    if  在册核对结果某单位sheet.cell(i+3,9).value is None:#如果核对表中上岗证为空
                        在册核对结果某单位sheet.cell(i+3,9).value = (item[0] + item[3] + item[4] + " " + chr(10)) if item[3] !="" else ""#复制上岗证
                    else:#如果核对表中上岗证不为空
                        在册核对结果某单位sheet.cell(i+3,9).value += (item[0] + item[3] + item[4] + " " + chr(10)) if item[3] !="" else ""#复制上岗证

for j in range(0,len(在册人员名单某单位姓名list)):#处理没有档案号的人员
    if 在册人员名单某单位姓名list[j][0]=="":
        在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row+1 ,5).value =  在册人员名单某单位姓名list[j][1]#在册姓名
        在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,6).value =  在册人员名单某单位姓名list[j][2]#在册身份证
        在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,10).value =  在册人员名单某单位姓名list[j][3]#在册职务
        for item in 井下所有单位姓名list:#item = [2, '综采一队', '吴雪峰', '142225197810234785']...
            if  在册人员名单某单位姓名list[j][2] in item:#如果身份证在item中,则把item中的档案号单位赋值到核对结果表中
                在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,7).value = str(item[0])+item[1]+chr(10)
                break
            elif 在册人员名单某单位姓名list[j][1] in item:#如果只有姓名在item中,则把item中的档案号单位身份证赋值到核对结果表中
                if  在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,7).value is None:
                    在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,7).value = str(item[0])+item[1]+ (item[3] if item[3] is not None else "")  + chr(10)
                else:
                    在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,7).value += str(item[0])+item[1]+ (item[3] if item[3] is not None else "")  + chr(10)
        for item in 持证表所有单位姓名list:
            if 在册人员名单某单位姓名list[j][2] in item:#item =  ['综采一队', '陈伟', '合格证所在单位', '井下单体支护工', '140211198212234739']
                在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,8).value = item[2]#如果身份证在item中,则把合格证所在单位赋值到核对结果表中
                在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,9).value = (item[3] if item[2] !="" else item[0] + item[3]) if item[3] != "" else "" #如果身份证在item中,则把工种赋值到核对结果表中
                break
            elif 在册人员名单某单位姓名list[j][1] in item:#如果身份证不在item中,而姓名在
                if  在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,8).value is None:#如果核对表中合格证所在单位为空
                    在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,8).value = (item[2] + item[4] + " " + chr(10)) if item[2] != "" else ""#把合格证所在单位与身份证都赋值到核对表中
                else:                                                                         #如果核对表中合格证所在单位不为空
                    在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,8).value += (item[2] + item[4] + " " + chr(10)) if item[2] != "" else ""
                if  在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,9).value is None:#如果核对表中上岗证为空
                    在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,9).value = (item[0] + item[3] + item[4] + " " + chr(10)) if item[3] !="" else ""#把上岗证与身份证赋值到核对表中
                else:#如果核对表中上岗证不为空
                    在册核对结果某单位sheet.cell(在册核对结果某单位sheet.max_row,9).value += (item[0] + item[3]  + item[4] + " " + chr(10)) if item[3] !="" else ""#复制上岗证

在册核对结果book.save("临时核对结果.xlsx")
# print(井下某单位姓名list)
# print(在册人员名单某单位姓名list)